import pandas as pd
import numpy as np
import os
import cv2
import math
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# Suppress OpenCV logs
os.environ["OPENCV_LOG_LEVEL"] = "SILENT"

# === Configuration ===
subset_folder = "/Users/chelseafowler/Documents/PhD_oysters/autosize/autosize2/length_Spring2025_tp1/Annotated_Images"
results_csv = "/Users/chelseafowler/Documents/PhD_oysters/autosize/autosize2/length_Spring2025_tp1/particle_analysis_results.csv"
tag_ids_file = "/Users/chelseafowler/Documents/PhD_oysters/autosize/autosize2/length_Spring2025_tp1/TagIDs050625.xlsx"
final_output_folder = "/Users/chelseafowler/Documents/PhD_oysters/autosize/autosize2/length_Spring2025_tp1/Final"
output_file = "/Users/chelseafowler/Documents/PhD_oysters/autosize/autosize2/length_Spring2025_tp1/final.xlsx"
warning_file = "/Users/chelseafowler/Documents/PhD_oysters/autosize/autosize2/length_Spring2025_tp1/warnings.txt"
calib_folder = "/Users/chelseafowler/Documents/PhD_oysters/autosize/autosize2/length_Spring2025_tp1/Calibration_Results"
# if column names in tagID is different, go to end of script and change keep columns
# Specify tag assignment reading order:
# 'landscape' = rows left-to-right, top-to-bottom (default)
# 'portrait'  = columns top-to-bottom, right-to-left
orientation = 'landscape'

os.makedirs(final_output_folder, exist_ok=True)

# === Warning logger ===
def log_warning(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(warning_file, "a") as wf:
        wf.write(f"[{ts}] {msg}\n")

# Initialize warnings file
txt = (
    "WARNING, PLEASE CHECK THE FOLLOWING FILES!\n"
    "PROBABLY TWO REASONS:\n"
    "1) Your TagID may be missing entries.\n"
    "2) The binary mask may have missing oysters.\n\n"
)
with open(warning_file, "w") as wf:
    wf.write(txt)

# === Load data ===
tag_ids = pd.read_excel(tag_ids_file, dtype={"date": str})
particle_results = pd.read_csv(results_csv)
particle_results["Label"] = particle_results["Label"].str.replace(".jpeg", "", regex=False)

# Ensure necessary columns exist
needed_cols = ["Farthest_X","Farthest_Y","Farthest_Length_mm","Red_X","Red_Y",
               "Area_mm2","Major_mm","Minor_mm","Angle"]
for col in needed_cols:
    if col not in tag_ids.columns:
        tag_ids[col] = np.nan

# Write initial TagIDs to Excel and load workbook
tag_ids.to_excel(output_file, index=False)
wb = load_workbook(output_file)
ws = wb.active

# === Ruler removal filter ===
def remove_ruler_by_area(group):
    # Drop the contour with the largest Area (the ruler)
    return group.drop(index=group['Area'].idxmax())

filtered = []
for label, grp in particle_results.groupby('Label'):
    expected = tag_ids[tag_ids['file_name']==label].shape[0]
    if grp.shape[0] == expected + 1:
        grp = remove_ruler_by_area(grp)
    filtered.append(grp)
particle_results = pd.concat(filtered, ignore_index=True)

# Excel highlight style
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# === Main processing loop ===
for label, grp in particle_results.groupby('Label'):
    mask = tag_ids['file_name'] == label
    tag_group = tag_ids[mask]
    # warn if mismatch
    if grp.shape[0] != tag_group.shape[0]:
        log_warning(f"{label}: {grp.shape[0]} blobs vs {tag_group.shape[0]} tags.")
        for r in tag_group.index:
            for cell in ws[r+2]: cell.fill = red_fill
        if grp.shape[0] > tag_group.shape[0]:
            grp = grp.iloc[:tag_group.shape[0]]
        else:
            tag_group = tag_group.iloc[:grp.shape[0]]

    # Sort into reading order based on orientation
    if orientation == 'landscape':
        # row-wise: top-to-bottom, left-to-right
        grp = grp.sort_values(['Y','X']).reset_index(drop=True)
        grp['row_grp'] = (grp['Y'].diff().fillna(999) > 150).cumsum()
        grp = grp.sort_values(['row_grp','X']).reset_index(drop=True)
    else:
        # column-wise: right-to-left, top-to-bottom
        grp = grp.sort_values(['X','Y'], ascending=[False, True]).reset_index(drop=True)
        grp['col_grp'] = (grp['X'].diff().abs().fillna(999) > 150).cumsum()
        grp = grp.sort_values(['col_grp','Y']).reset_index(drop=True)

    # Load calibration\    
    calib_file = os.path.join(calib_folder, f'calibration_{label}.txt')
    try:
        with open(calib_file) as f:
            pixel_mm = float(f.readline().split(':',1)[1].strip())
    except Exception as e:
        log_warning(f"{label}: cannot load calibration ({e}), default px/mm=1")
        pixel_mm = 1.0

    # Assign numbers and compute mm measurements
    grp['number'] = tag_group['number'].values
    for idx in tag_group.index:
        pr = grp[grp['number']==tag_group.at[idx,'number']].iloc[0]
        # pixel measurements
        rx, ry = pr['Red_X'], pr['Red_Y']
        fx, fy = pr['Farthest_X'], pr['Farthest_Y']
        # convert to mm
        tag_ids.at[idx, 'Area_mm2'] = pr['Area'] / (pixel_mm**2)
        tag_ids.at[idx, 'Major_mm']   = pr['Major'] / pixel_mm
        tag_ids.at[idx, 'Minor_mm']   = pr['Minor'] / pixel_mm
        tag_ids.at[idx, 'Farthest_Length_mm'] = math.hypot(fx-rx, fy-ry) / pixel_mm
        tag_ids.at[idx, 'Red_X']      = rx
        tag_ids.at[idx, 'Red_Y']      = ry
        tag_ids.at[idx, 'Farthest_X'] = fx
        tag_ids.at[idx, 'Farthest_Y'] = fy
        tag_ids.at[idx, 'Angle']      = pr['Angle']

    # Draw on image
    img_path = os.path.join(subset_folder, f"{label}.jpeg")
    img = cv2.imread(img_path)
    if img is None:
        img = cv2.imread(img_path.replace('.jpeg','.jpg'))
    if img is None:
        log_warning(f"{label}: cannot load image file.")
        continue
    out = img.copy()
    for _, pr in grp.iterrows():
        x, y = int(pr['X']), int(pr['Y'])
        a, m, ang = pr['Major'], pr['Minor'], math.radians(pr['Angle'])
        dx, dy = (a/2)*math.cos(ang), (a/2)*math.sin(ang)
        cv2.line(out, (x-int(dx), y-int(dy)), (x+int(dx), y+int(dy)), (0,255,128), 2)
        dx2, dy2 = (m/2)*math.cos(ang+math.pi/2), (m/2)*math.sin(ang+math.pi/2)
        cv2.line(out, (x-int(dx2), y-int(dy2)), (x+int(dx2), y+int(dy2)), (0,255,128), 2)
        cv2.putText(out, str(pr['number']), (x, y+int(m/2)-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,128), 2)
        # draw red line
        if not (pd.isna(pr['Red_X']) or pd.isna(pr['Farthest_X'])):
            cv2.line(out, (int(pr['Red_X']), int(pr['Red_Y'])), (int(pr['Farthest_X']), int(pr['Farthest_Y'])), (0,0,255), 2)
    cv2.imwrite(os.path.join(final_output_folder, f"Final_{label}.jpeg"), out)

# === Final Excel writeback ===
keep_cols = ["date","file_name","tag","number","rep","treatment","ploidy","status","experiment","week",
             "Farthest_X","Farthest_Y","Farthest_Length_mm",
             "Red_X","Red_Y","Area_mm2","Major_mm","Minor_mm","Angle"]
tag_ids = tag_ids[keep_cols]
for i, row in tag_ids.iterrows():
    if pd.isna(row['Red_X']) or pd.isna(row['Farthest_X']):
        for cell in ws[i+2]: cell.fill = red_fill
# write values
for r_i, row in tag_ids.iterrows():
    for c_i, val in enumerate(row, 1): ws.cell(row=r_i+2, column=c_i, value=val)
wb.save(output_file)
print(f"Updated data saved to {output_file}. Warnings logged to {warning_file}.")
